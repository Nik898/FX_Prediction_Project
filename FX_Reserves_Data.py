from airflow import DAG
from datetime import timedelta, datetime
from airflow.utils.dates import days_ago
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
from airflow.operators.email import EmailOperator
import requests
from requests import get
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import openpyxl
from datetime import date
from datetime import timedelta
from datetime import datetime
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import xlsxwriter
import re
import pandas as pd
import os
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import sys
import json
import pytz

def extract_fxres(**kwargs):
    link =[]
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    url1 = "https://www.rbi.org.in/Scripts/WSSViewDetail.aspx?TYPE=Section&PARAM1=2"
    results1 = requests.get(url1, headers=headers, verify=False) 
    soup = BeautifulSoup(results1.text, "html.parser")
    table = soup.find('table',{"class":"tablebg"})
    links = table.find_all('a',{"class": "link2"})
    for i in links:
        link.append(str(i))

    return link

def transform_fxres(**context):
    urls =[]
    links = list(context['ti'].xcom_pull(task_ids='extract'))
    for i in links:
        url = str(i)
        url = url[url.find('href="')+6:]
        url = url[:url.find('"')]
        furl = 'https://www.rbi.org.in/Scripts/' + url
        urls.append(furl)

    return urls

def load_fxres(**context):
    links = list(context['ti'].xcom_pull(task_ids='transform'))
    wb = load_workbook(filename='/root/airflow/files/links2.xlsx')
    sheet = wb.worksheets[2]
    sheet_a =[]
    for cell in sheet["A"]:
        sheet_a.append(cell.value)
    for i in range(len(links)):
        if links[i] in sheet_a:
            continue
        else:
            index_a ='A'+str(sheet.max_row+1)
            sheet[index_a]= links[i]
            sheet_a.append(links[i])
    wb.close()
    wb.save(filename='/root/airflow/files/links2.xlsx')

def scrape(url):
    try:
        headers = {"Accept-Language": "en-US, en;q=0.5"}

        results1 = requests.get(url, headers=headers, verify=True) 
        soup = BeautifulSoup(results1.text, "html.parser")

        table_arr = soup.find_all('table', {'class' : 'tablebg'})
        elem_arr = table_arr[0].find_all('tr')

        date_elem = str(elem_arr[1])
        date_elem = date_elem[date_elem.find("</b>")+4:]
        date_elem = date_elem[:date_elem.find("</td>")]
        date_elem = date_elem.replace(",","")
        date_elem = " ".join(date_elem.strip().split())
        end_date = datetime.strptime(date_elem, '%b %d %Y')

        elem_arr = table_arr[1].find_all('tr')

        date_elem = str(elem_arr[0])
        date_elem = date_elem[date_elem.find("As on")+6:]
        date_elem = date_elem[:date_elem.find("</span")]
        date_elem = date_elem.replace(",","")
        date_elem = " ".join(date_elem.strip().split())
        start_date = datetime.strptime(date_elem, '%B %d %Y')

        tot_res_str = str(elem_arr[4].find_all('td')[2])
        tot_res_str = tot_res_str[tot_res_str.find('>')+1:]
        tot_res_str = tot_res_str[:tot_res_str.find('<')]
        tot_res_str = tot_res_str.replace(",","")
        try:
            tot_reserve = float(tot_res_str)
            tot_reserve = tot_reserve/1000
        except:
            tot_reserve = None

        for_cur_str = str(elem_arr[5].find_all('td')[2])
        for_cur_str = for_cur_str[for_cur_str.find('>')+1:]
        for_cur_str = for_cur_str[:for_cur_str.find('<')]
        for_cur_str = for_cur_str.replace(",","")
        try:
            foreign_currency = float(for_cur_str)
            foreign_currency = foreign_currency/1000
        except:
            foreign_currency = None

        gold_str = str(elem_arr[6].find_all('td')[2])
        gold_str = gold_str[gold_str.find('>')+1:]
        gold_str = gold_str[:gold_str.find('<')]
        gold_str = gold_str.replace(",","")
        try:
            gold = float(gold_str)
            gold = gold/1000
        except:
            gold = None

        return start_date, end_date, tot_reserve, foreign_currency, gold
    
    except: 
        return 0

def extract(**kwargs):
    start =[]
    end =[]
    tot_reserve =[]
    foreign_currency=[]
    gold =[]
    df = pd.read_excel('/root/airflow/files/links2.xlsx',engine='openpyxl', sheet_name='FX_Reserves')
    links = df['Urls'].tolist()
    for i in links:
        if scrape(i) ==0:
            continue
        else:
            sd,ed,tr,fc,g = scrape(i)
            start.append(sd)
            end.append(ed)
            tot_reserve.append(tr)
            foreign_currency.append(fc)
            gold.append(g)

    return start,end,tot_reserve,foreign_currency,gold

def load(**context):
    start = list(context['ti'].xcom_pull(task_ids='extract_data')[0])
    end = list(context['ti'].xcom_pull(task_ids='extract_data')[1])
    tot_reserve = list(context['ti'].xcom_pull(task_ids='extract_data')[2])
    foreign_currency = list(context['ti'].xcom_pull(task_ids='extract_data')[3])
    gold = list(context['ti'].xcom_pull(task_ids='extract_data')[4])

    for k in range(len(start)):
        sd = start[k]
        ed = end[k]
        tot = tot_reserve[k]
        fc = foreign_currency[k]
        g = gold[k]

        wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
        sheet1 = wb.worksheets[1]
        dates =[]
        for c in sheet1['A']:
            dates.append(c.value) 
        i = sd
        check_date = "14-02-2020"
        check_date = datetime.strptime(check_date, "%d-%m-%Y")
        print(f"Start = {i}")
        print(f"End = {ed}")
        while (i <= ed):
            if check_date.date() > i.date():
                i += timedelta(days=1)
                continue
            x = i.strftime("%d-%m-%Y")
            print(f"Current ={x}")
            if x in dates:
                for cell in sheet1['A']:
                    if cell.value == x:
                        #print("y")
                        #print(cell.row)
                        sheet1["AD" + str(cell.row)] = tot
                        #d2 = "C" + str(cell.row)
                        #print(sheet1[d2].value)
                        sheet1["AE" + str(cell.row)] = fc
                        sheet1["AF" + str(cell.row)] = g
                        break
            else:
                row = sheet1.max_row+1
                d1 = "A" + str(row)
                d2 = "AD" + str(row)
                d3 = "AE" + str(row)
                d4 = "AF" + str(row)

                sheet1["A" + str(row)] = x
                print(sheet1[d1].value)
                sheet1["AD" + str(row)] = tot
                print(sheet1[d2].value)
                sheet1["AE" + str(row)] = fc
                print(sheet1[d3].value)
                sheet1["AF" + str(row)] = g
                print(sheet1[d4].value)
                dates.append(x)

            i += timedelta(days=1)

        wb.close()
        wb.save(filename='/root/airflow/files/Data.xlsx')
    return
    

dt = datetime.now()

default_args = {
    'owner': 'airflow',
    'depends_on_past': False,
    'start_date': datetime(2023,1,4),
    'email': ['nikheleshbhattacharyya@gmail.com'],
    'email_on_failure': True,
    'email_on_retry': False,
    'retries': 3,
    'retry_delay': timedelta(minutes=2)
}

schedule_interval = '0 */2 * * *'

dag = DAG(
    dag_id = 'FX_Reserves',
    default_args = default_args,
    schedule = '@daily'
)

task_start = BashOperator(
    task_id = 'start_task',
    bash_command = 'echo start',
    dag = dag
)

task1 = PythonOperator(
    task_id = 'extract',
    python_callable = extract_fxres,
    provide_context = True,
    dag = dag
)
# Task 2: save the daily summary data
task2 = PythonOperator(
    task_id = 'transform',
    python_callable = transform_fxres,
    provide_context = True,
    dag = dag
)
# Task 3: scraping daily provinces data
task3 = PythonOperator(
    task_id = 'load',
    python_callable = load_fxres,
    provide_context = True,
    dag = dag
)

task4 = PythonOperator(
    task_id = 'extract_data',
    python_callable = extract,
    provide_context = True,
    dag = dag
)

task5 = PythonOperator(
    task_id = 'load_data',
    python_callable = load,
    provide_context = True,
    dag = dag
)


#send_email = EmailOperator(
#    task_id = 'send_email',
#    to = ['nikheleshbhattacharyya@gmail.com'],
#    subject = 'Covid19 data ',
#    html_content = '''
#    {date} Data data has been scraped!
#    '''.format(date = datetime.now()),
#    dag = dag
#)

finish_start = BashOperator(
    task_id = 'finish_task',
    bash_command = 'echo finish',
    dag = dag
)

task_start>>task1>>task2>>task3>>task4>>task5>>finish_start