from airflow import DAG
from datetime import timedelta, datetime
from airflow.utils.dates import days_ago
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
from airflow.operators.email import EmailOperator
import requests
from requests import get
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import openpyxl
from datetime import date
from datetime import timedelta
from datetime import datetime
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import xlsxwriter
import re
import pandas as pd
import os
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import sys
import json
import pytz
#print(sys.setrecursionlimit(4000))

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument('--disable-dev-shm-usage')
options = Options()

options.headless = True

def change_dates(dates,start,end):
    act_dates = []
    for date in dates:
        try:
            d1 = datetime.strptime(date,"%m-%d-%Y")
        except:
            try:
                d1 = datetime.strptime(date,"%m-%d-%y")
            except:
                try:
                    d1 = datetime.strptime(date,"%b-%d-%Y")
                except:
                    try:
                        d1 = datetime.strptime(date,"%B-%d-%Y")
                    except:
                        d1 = None

        try:
            d2 = datetime.strptime(date,"%d-%m-%Y")
        except:
            try:
                d2 = datetime.strptime(date,"%d-%m-%y")
            except:
                try:
                    d2 = datetime.strptime(date,"%d-%b-%Y")
                except:
                    try:
                        d2 = datetime.strptime(date,"%d-%B-%Y")
                    except:
                        d2 = None

        if d1 == None:
            actual_date = d2
            act_dates.append(d2)
        elif d2 == None:
            actual_date = d1
            act_dates.append(d1)
        else:
            if start<=d1<=end:
                actual_date = d1
                act_dates.append(d1)
            else:
                actual_date = d2
                act_dates.append(d2)

    return act_dates

def check_dates(data):
    check_date = data[data.find("Turnover Data")+15:data.find("The")].strip()
    try:
        year = check_date[check_date.find(",")+1:].strip()
        start = check_date[:check_date.find("-")].strip()
        end = check_date[check_date.find("-")+1:].strip()
        end = end[:end.find(",")]
        if start.find("December")!= -1 and end.find("January")!= -1:
            year2 = int(year)
            year2= year2 -1
            year2 = str(year2)
            start = start+" "+year2
            end = end+" "+year
            start = datetime.strptime(start,"%B %d %Y")
            end = datetime.strptime(end,"%B %d %Y")
        else:
            start = start+" "+year
            end = end+" "+year
            start = datetime.strptime(start,"%B %d %Y")
            end = datetime.strptime(end,"%B %d %Y")

    except:
        year = check_date[check_date.find(",")+1:].strip()
        start = check_date[:check_date.find("–")].strip()
        end = check_date[check_date.find("–")+1:].strip()
        end = end[:end.find(",")]
        if start.find("December")!= -1 and end.find("January")!= -1:
            year2 = int(year)
            year2= year2 -1
            year2 = str(year2)
            start = start+" "+year2
            end = end+" "+year
            start = datetime.strptime(start,"%B %d %Y")
            end = datetime.strptime(end,"%B %d %Y")
        else:
            start = start+" "+year
            end = end+" "+year
            start = datetime.strptime(start,"%B %d %Y")
            end = datetime.strptime(end,"%B %d %Y")

    return start,end

def extract_fxturn(**kwargs):
    url = "https://www.rbi.org.in/Scripts/BS_ForeignExchangeDisplay.aspx"
    driver = webdriver.Chrome('/usr/local/bin/chromedriver',options=chrome_options)
    driver.implicitly_wait(30)
    driver.get(url)
    python_button1 = driver.find_elements(By.ID, 'divArchiveMain')
    python_button1[0].click()
    urls=[]
    soup_level1=BeautifulSoup(driver.page_source, 'lxml')
    for i in range(2020,2024):
        id1 = "btn"+str(i)
        id2 = str(i)+"0"
        python_button2 = driver.find_elements(By.ID, id1)
        python_button2[0].click()
        soup_level2=BeautifulSoup(driver.page_source, 'lxml')
        python_button3 = driver.find_elements(By.ID, id2)
        python_button3[0].click()
        soup_level3=BeautifulSoup(driver.page_source, 'lxml')
        link = soup_level3.find_all('a',{"class": "link2"})
        for i in link:
            urls.append(str(i))
    return urls

def transform_fxturn(**context):
    link = []
    urls = list(context['ti'].xcom_pull(task_ids='extract'))
    for ll in urls:
        urlk = str(ll)[str(ll).find("href=")+6:]
        urlk = urlk[:urlk.find('"')]
        urlk = "https://www.rbi.org.in/Scripts/" + urlk
        lk = str(ll)
        date = lk[lk.find("Money Market Operations as on")+30:lk.find("</a>")]
        #date = date.replace(",","")
        #if date.startswith('wMMO') == True:
        #    continue
        link.append(urlk)
        #date_time.append(date)
    return link


def load_fxturn(**context):
    links = list(context['ti'].xcom_pull(task_ids='transform'))
    wb = load_workbook(filename='/root/airflow/files/links2.xlsx')
    sheet = wb.worksheets[1]
    sheet_a =[]
    for cell in sheet["A"]:
        sheet_a.append(cell.value)
    for i in range(len(links)):
        if links[i] in sheet_a:
            continue
        else:
            index_a ='A'+str(sheet.max_row+1)
            sheet[index_a]= links[i]
            sheet_a.append(links[i])
    wb.close()
    wb.save(filename='/root/airflow/files/links2.xlsx')
    return

def extract_data(**kwargs):
    df = pd.read_excel('/root/airflow/files/links2.xlsx',engine='openpyxl', sheet_name='FX_Turnover')
    urls = df['Urls'].tolist()
    error = []
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    dd1 =[]
    dd2 =[]
    purchases =[]
    sales = []
    for url in urls:
        try:
            #print(url)
            results = requests.get(url, headers=headers, verify=True) 
            soup = BeautifulSoup(results.text, "html.parser")
            running2 = soup.find_all('table', class_="tablebg")

            data = running2[0].text
            data = " ".join(data.strip().split())
            start,end =check_dates(data)
            str1 ="Purchases"
            str2 ="Purchase"
            if str1 in data:
                pur = data[data.find("Purchases")+10:data.find("Sales")-1]
            else:
                pur = data[data.find("Purchase")+9:data.find("Sales")-1]
            pur = pur.replace("/","-")
            pur = pur.replace(",","")
            pur = pur.split(" ")
            n=13
            pur = [' '.join(pur[i:i+n]) for i in range(0,len(pur),n)]
            dates1 =[]
            for p in pur:
                p = p.split(" ")
                dates1.append(p[0])
                purchases.append(p[1:13])
            d1 = change_dates(dates1,start,end)
            dd1+=d1
            sal = data[data.find("Sales")+5:data.find("(Provisional Data)")].strip()
            sal = sal.replace("/","-")
            sal = sal.replace(",","")
            sal = sal.split(" ")
            n=13
            sal = [' '.join(sal[i:i+n]) for i in range(0,len(sal),n)]
            dates2 =[]
            for p in sal:
                p = p.split(" ")
                dates2.append(p[0])
                sales.append(p[1:13])
            d2 = change_dates(dates2,start,end)
            dd2+=d2
        except:
            #print(url)
            error.append(url)
            continue

    return dd1,dd2,purchases,sales

def transform_data(**context):
    date1 = list(context['ti'].xcom_pull(task_ids='extract_data')[0])
    date2 = list(context['ti'].xcom_pull(task_ids='extract_data')[1])
    p = list(context['ti'].xcom_pull(task_ids='extract_data')[2])
    s = list(context['ti'].xcom_pull(task_ids='extract_data')[3])
    dat1 = []
    dat2 = []
    for d1 in date1:
        x1 = d1.strftime('%d-%m-%Y')
        dat1.append(x1)
    for d2 in date2:
        x2 = d2.strftime('%d-%m-%Y')
        dat2.append(x2)

    return dat1,dat2,p,s

def load_data(**context):
    date1 = list(context['ti'].xcom_pull(task_ids='transform_data')[0])
    date2 = list(context['ti'].xcom_pull(task_ids='transform_data')[1])
    pur = list(context['ti'].xcom_pull(task_ids='transform_data')[2])
    sal = list(context['ti'].xcom_pull(task_ids='transform_data')[3])

    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    check_date = "14-02-2020"
    check_date = datetime.strptime(check_date, "%d-%m-%Y")
    sheet = wb.worksheets[1]
    sheet_a =[]
    for cell in sheet["A"]:
        sheet_a.append(cell.value)
    for i in range(len(date1)):
        x = date1[i]
        p = pur[i] 
        y = datetime.strptime(x, "%d-%m-%Y")
        if y.date()<check_date.date():
            continue
        if x in sheet_a:
                for cell in sheet['A']:
                    if cell.value == x:
                        sheet["F" + str(cell.row)] = float(p[0])
                        sheet["G" + str(cell.row)] = float(p[1])
                        sheet["H" + str(cell.row)] = float(p[2])
                        sheet["I" + str(cell.row)] = float(p[3])
                        sheet["J" + str(cell.row)] = float(p[4])
                        sheet["K" + str(cell.row)] = float(p[5])
                        sheet["L" + str(cell.row)] = float(p[6])
                        sheet["M" + str(cell.row)] = float(p[7])
                        sheet["N" + str(cell.row)] = float(p[8])
                        sheet["O" + str(cell.row)] = float(p[9])
                        sheet["P" + str(cell.row)] = float(p[10])
                        sheet["Q" + str(cell.row)] = float(p[11])
                        break
        else:
            row = sheet.max_row+1
            sheet["A" + str(row)] = x
            sheet["F" + str(row)] = float(p[0])
            sheet["G" + str(row)] = float(p[1])
            sheet["H" + str(row)] = float(p[2])
            sheet["I" + str(row)] = float(p[3])
            sheet["J" + str(row)] = float(p[4])
            sheet["K" + str(row)] = float(p[5])
            sheet["L" + str(row)] = float(p[6])
            sheet["M" + str(row)] = float(p[7])
            sheet["N" + str(row)] = float(p[8])
            sheet["O" + str(row)] = float(p[9])
            sheet["P" + str(row)] = float(p[10])
            sheet["Q" + str(row)] = float(p[11])
            sheet_a.append(x)
    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')

    for i in range(len(date2)):
        x = date2[i]
        s = sal[i] 
        y = datetime.strptime(x, "%d-%m-%Y")
        if y.date()<check_date.date():
            continue
        if x in sheet_a:
                for cell in sheet['A']:
                    if cell.value == x:
                        sheet["R" + str(cell.row)] = float(s[0])
                        sheet["S" + str(cell.row)] = float(s[1])
                        sheet["T" + str(cell.row)] = float(s[2])
                        sheet["U" + str(cell.row)] = float(s[3])
                        sheet["V" + str(cell.row)] = float(s[4])
                        sheet["W" + str(cell.row)] = float(s[5])
                        sheet["X" + str(cell.row)] = float(s[6])
                        sheet["Y" + str(cell.row)] = float(s[7])
                        sheet["Z" + str(cell.row)] = float(s[8])
                        sheet["AA" + str(cell.row)] = float(s[9])
                        sheet["AB" + str(cell.row)] = float(s[10])
                        sheet["AC" + str(cell.row)] = float(s[11])
                        break
        else:
            row = sheet.max_row+1
            sheet["A" + str(row)] = x
            sheet["R" + str(row)] = float(s[0])
            sheet["S" + str(row)] = float(s[1])
            sheet["T" + str(row)] = float(s[2])
            sheet["U" + str(row)] = float(s[3])
            sheet["V" + str(row)] = float(s[4])
            sheet["W" + str(row)] = float(s[5])
            sheet["X" + str(row)] = float(s[6])
            sheet["Y" + str(row)] = float(s[7])
            sheet["Z" + str(row)] = float(s[8])
            sheet["AA" + str(row)] = float(s[9])
            sheet["AB" + str(row)] = float(s[10])
            sheet["AC" + str(row)] = float(s[11])
            sheet_a.append(x)
    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')


dt = datetime.now()

default_args = {
    'owner': 'airflow',
    'depends_on_past': False,
    'start_date': datetime(2023,1,4),
    'email': ['nikheleshbhattacharyya@gmail.com'],
    'email_on_failure': True,
    'email_on_retry': False,
    'retries': 3,
    'retry_delay': timedelta(minutes=2)
}

schedule_interval = '0 */2 * * *'

dag = DAG(
    dag_id = 'FX__Turnover',
    default_args = default_args,
    schedule = '@daily'
)

task_start = BashOperator(
    task_id = 'start_task',
    bash_command = 'echo start',
    dag = dag
)

task1 = PythonOperator(
    task_id = 'extract',
    python_callable = extract_fxturn,
    provide_context = True,
    dag = dag
)

task2 = PythonOperator(
    task_id = 'transform',
    python_callable = transform_fxturn,
    provide_context = True,
    dag = dag
)

task3 = PythonOperator(
    task_id = 'load',
    python_callable = load_fxturn,
    provide_context = True,
    dag = dag
)

task4 = PythonOperator(
    task_id = 'extract_data',
    python_callable = extract_data,
    provide_context = True,
    dag = dag
)

task5 = PythonOperator(
    task_id = 'transform_data',
    python_callable = transform_data,
    provide_context = True,
    dag = dag
)

task6 = PythonOperator(
    task_id = 'load_data',
    python_callable = load_data,
    provide_context = True,
    dag = dag
)

#send_email = EmailOperator(
#    task_id = 'send_email',
#    to = ['nikheleshbhattacharyya@gmail.com'],
#    subject = 'Covid19 data ',
#    html_content = '''
#    {date} Data data has been scraped!
#    '''.format(date = datetime.now()),
#    dag = dag
#)

finish_start = BashOperator(
    task_id = 'finish_task',
    bash_command = 'echo finish',
    dag = dag
)

task_start>>task1>>task2>>task3>>task4>>task5>>task6>>finish_start