from airflow import DAG
from datetime import timedelta, datetime
from airflow.utils.dates import days_ago
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
from airflow.operators.email import EmailOperator
import requests
from requests import get
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import openpyxl
from datetime import date
from datetime import timedelta
from datetime import datetime
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import xlsxwriter
import re
import pandas as pd
import os
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import sys
import json
import pytz
#print(sys.setrecursionlimit(4000))

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument('--disable-dev-shm-usage')
options = Options()

options.headless = True


def extract_MM(**kwargs):
    url = "https://www.rbi.org.in/Scripts/BS_ViewMMO.aspx"
    driver = webdriver.Chrome('/usr/local/bin/chromedriver',options=chrome_options)
    driver.implicitly_wait(30)
    driver.get(url)
    python_button1 = driver.find_elements(By.ID, 'divArchiveMain')
    python_button1[0].click()
    urls=[]
    soup_level1=BeautifulSoup(driver.page_source, 'lxml')
    for i in range(2018,2024):
        id1 = "btn"+str(i)
        id2 = str(i)+"0"
        python_button2 = driver.find_elements(By.ID, id1)
        python_button2[0].click()
        soup_level2=BeautifulSoup(driver.page_source, 'lxml')
        python_button3 = driver.find_elements(By.ID, id2)
        python_button3[0].click()
        soup_level3=BeautifulSoup(driver.page_source, 'lxml')
        link = soup_level3.find_all('a',{"class": "link2"})
        for i in link:
            urls.append(str(i))
    
    #with open("extract.json", "w") as fp:
    #    json.dump(urls, fp)

    return urls


def transform_MM(**context):
    link = []
    date_time = []
    #with open('extract.json', 'rb') as fp:
        #urls = json.load(fp)
    urls = list(context['ti'].xcom_pull(task_ids='extract_link'))
    for ll in urls:
        urlk = str(ll)[str(ll).find("href=")+6:]
        urlk = urlk[:urlk.find('"')]
        urlk = "https://www.rbi.org.in/Scripts/" + urlk
        lk = str(ll)
        date = lk[lk.find("Money Market Operations as on")+30:lk.find("</a>")]
        date = date.replace(",","")
        if date.startswith('wMMO') == True:
            continue
        link.append(urlk)
        date_time.append(date)
    
    #with open("transform_link.json", "w") as fp:
    #    json.dump(link, fp)

    #with open("transform_date.json", "w") as fp:
    #    json.dump(date_time, fp)

    return link,date_time

def load_MM(**context):
    urls = list(context['ti'].xcom_pull(task_ids='transform_link')[0])
    dates = list(context['ti'].xcom_pull(task_ids='transform_link')[1])
    #with open('transform_link.json', 'rb') as fp:
    #    urls = json.load(fp)
    #with open('transform_date.json', 'rb') as fp:
    #    dates = json.load(fp)
    wb = load_workbook(filename='/root/airflow/files/links2.xlsx')
    sheet = wb.worksheets[0]
    sheet_a =[]
    c=0
    for cell in sheet["A"]:
        sheet_a.append(cell.value)
    for i in range(len(dates)):
        if dates[i] in sheet_a:
            continue
        else:
            c=c+1
            print(c)
            index_a ='A'+str(sheet.max_row+1)
            index_b = 'B'+str(sheet.max_row+1)
            sheet[index_a]= dates[i]
            sheet[index_b]= urls[i]
            sheet_a.append(dates[i])
    wb.close()
    wb.save(filename='/root/airflow/files/links2.xlsx')
    return

def scrape_RBI_MM(url1):
    try:
        headers = {"Accept-Language": "en-US, en;q=0.5"}
        results1 = requests.get(url1, headers=headers, verify=True) 
        soup1 = BeautifulSoup(results1.text, "html.parser")

        running = soup1.find_all('tr', class_="tableContent1")
        running2 = soup1.find_all('td', class_="tableheader")

        running = soup1.find_all('tr', class_="tableContent1")
        #running2 = soup1.find_all('td', class_="tableheader")
        rs = running[0].find_all('td')
        lists = str(rs[0].text)
        substr = "Net liquidity injected (outstanding including today's operations) [injection (+)/absorption (-)]*"
        substr2 = "Net liquidity injected (outstanding including today's<strong> </strong>operations) [injection (+)/absorption (-)]*"
        if substr in lists:
            val = lists[lists.find(substr)+97:lists.find("RESERVE POSITION@")]
            val = val.replace("\xa0","")
            val = " ".join(val.strip().split())
            val = str(val)
            val = val.replace(",","")
            RBI_MM = float(val)
            RBI_MMx= RBI_MM/100000
        elif substr2 in lists:
            val = lists[lists.find(substr2)+114:lists.find("RESERVE POSITION@")]
            val = val.replace("\xa0","")
            val = " ".join(val.strip().split())
            val = str(val)
            val = val.replace(",","")
            RBI_MM = float(val)
            RBI_MMx= RBI_MM/100000

        for i in running2:
            substr2 = 'Money Market Operations as on'
            j = str(i)
            if substr2 in j:
                j = j[j.find('as on') + 6:]
                j = j[:j.find('</b>')]
                if j[-4] != "2": j = j[:j.find('(')-1]
                j = j.replace(",","")
                RBI_MM_date = datetime.strptime(j, '%B %d %Y')
                print(RBI_MM_date)

        running3 = soup1.find_all('table', class_="brd-ptable1")
        s = running3[0].text
        s = s.replace("\xa0","")
        s = " ".join(s.strip().split())
        call_money = s[s.find("Call Money")+11:]
        call_money = call_money.split(" ")
        try:
            call_money = float(call_money[1])
        except:
            call_money = None

        triparty_money = s[s.find("Triparty Repo")+14:]
        triparty_money = triparty_money.split(" ")
        try:
            triparty_money = float(triparty_money[1])
        except:
            triparty_money = None

        market_repo = s[s.find("Market Repo")+12:]
        market_repo = market_repo.split(" ")
        try:
            market_repo = float(market_repo[1])
        except:
            market_repo = None

        return RBI_MM_date, RBI_MMx, call_money,triparty_money,market_repo
    except: 
        print("error")
        return 0

def extract_mm(**kwargs):
    df = pd.read_excel('/root/airflow/files/links2.xlsx',engine='openpyxl', sheet_name='Money_Market')
    links = df['Urls'].tolist()
    Rbi_mm =[]
    Rbi_date =[]
    cm = []
    tm =[]
    mr =[]
    c=0
    for i in links:
        if scrape_RBI_MM(i) == 0:
            c=c+1
            print(c)
            continue
        else:
            RBI_MM_date,RBI_MM,call_money,triparty_money,market_repo = scrape_RBI_MM(i)
            Rbi_mm.append(RBI_MM)
            Rbi_date.append(RBI_MM_date)
            cm.append(call_money)
            tm.append(triparty_money)
            mr.append(market_repo)

    return Rbi_date,Rbi_mm,cm,tm,mr

def transform_mm(**context):
    values = list(context['ti'].xcom_pull(task_ids='extract_data')[1])
    dates = list(context['ti'].xcom_pull(task_ids='extract_data')[0])
    cm = list(context['ti'].xcom_pull(task_ids='extract_data')[2])
    tm = list(context['ti'].xcom_pull(task_ids='extract_data')[3])
    mr = list(context['ti'].xcom_pull(task_ids='extract_data')[4])
    
    dat = []
    for d in dates:
        x = d.strftime('%d-%m-%Y')
        dat.append(x)

    return values,dat,cm,tm,mr

def load_mm(**context):
    values = list(context['ti'].xcom_pull(task_ids='transform_data')[0])
    dates = list(context['ti'].xcom_pull(task_ids='transform_data')[1])
    cm = list(context['ti'].xcom_pull(task_ids='transform_data')[2])
    tm = list(context['ti'].xcom_pull(task_ids='transform_data')[3])
    mr = list(context['ti'].xcom_pull(task_ids='transform_data')[4])
    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    sheet = wb.worksheets[1]
    sheet_a =[]
    sheet_b =[]
    c=0
    for cell in sheet["A"]:
        sheet_a.append(cell.value)
    for cell in sheet["B"]:
        sheet_b.append(cell.value)
    for i in range(len(dates)):
        x = dates[i]
        if x in sheet_a:
                for cell in sheet['A']:
                    if cell.value == x:
                        #print("y")
                        #print(cell.row)
                        sheet["B" + str(cell.row)] = values[i]
                        sheet["C" + str(cell.row)] = cm[i]
                        sheet["D" + str(cell.row)] = tm[i]
                        sheet["E" + str(cell.row)] = mr[i]
                        break
        else:
            row = sheet.max_row+1
            d1 = "A" + str(row)
            d2 = "B" + str(row)
            d3 = "C" + str(row)
            d4 = "D" + str(row)
            d5 = "E" + str(row)

            sheet["A" + str(row)] = x
            print(sheet[d1].value)
            sheet["B" + str(row)] = values[i]
            print(sheet[d2].value)
            sheet["C" + str(row)] = cm[i]
            print(sheet[d3].value)
            sheet["D" + str(row)] = tm[i]
            print(sheet[d4].value)
            sheet["E" + str(row)] = mr[i]
            print(sheet[d5].value)
            dates.append(x)
    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')


dt = datetime.now()

default_args = {
    'owner': 'airflow',
    'depends_on_past': False,
    'start_date': datetime(2022,12,30),
    'email': ['nikheleshbhattacharyya@gmail.com'],
    'email_on_failure': True,
    'email_on_retry': False,
    'retries': 3,
    'retry_delay': timedelta(minutes=2)
}

schedule_interval = '0 */2 * * *'

dag = DAG(
    dag_id = 'Money_Market',
    default_args = default_args,
    schedule = '@daily'
)

task_start = BashOperator(
    task_id = 'start_task',
    bash_command = 'echo start',
    dag = dag
)

extract_link = PythonOperator(
    task_id = 'extract_link',
    python_callable = extract_MM,
    provide_context = True,
    dag = dag
)
# Task 2: save the daily summary data
transform_link = PythonOperator(
    task_id = 'transform_link',
    python_callable = transform_MM,
    provide_context = True,
    dag = dag
)
# Task 3: scraping daily provinces data
load_link = PythonOperator(
    task_id = 'load_link',
    python_callable = load_MM,
    provide_context = True,
    dag = dag
)

extract = PythonOperator(
    task_id = 'extract_data',
    python_callable = extract_mm,
    provide_context = True,
    dag = dag
)

transform = PythonOperator(
    task_id = 'transform_data',
    python_callable = transform_mm,
    provide_context = True,
    dag = dag
)

load = PythonOperator(
    task_id = 'load_data',
    python_callable = load_mm,
    provide_context = True,
    dag = dag
)

#send_email = EmailOperator(
#    task_id = 'send_email',
#    to = ['nikheleshbhattacharyya@gmail.com'],
#    subject = 'Covid19 data ',
#    html_content = '''
#    {date} Data data has been scraped!
#    '''.format(date = datetime.now()),
#    dag = dag
#)

finish_start = BashOperator(
    task_id = 'finish_task',
    bash_command = 'echo finish',
    dag = dag
)

task_start>>extract_link>>transform_link>>load_link>>extract>>transform>>load>>finish_start