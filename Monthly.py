from airflow import DAG
from datetime import timedelta, datetime
from airflow.utils.dates import days_ago
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
from airflow.operators.email import EmailOperator
import requests
from requests import get
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import openpyxl
from datetime import date
from datetime import timedelta
from datetime import datetime
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import xlsxwriter
import re
import pandas as pd
import os
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import sys
import json
import pytz

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument('--disable-dev-shm-usage')
options = Options()

options.headless = True

def link_gen(elem):
    bulletin_4 = []
    bulletin_4A = []
    bulletin_31 = []
    bulletin_34 = []
    bulletin_37 = []
    bulletin_35 = []
    bulletin_36 = []
    for a in elem.find_elements(By.TAG_NAME,"tr"):
        i = a.get_attribute('innerHTML')
        if "4. Sale/Purchase of U.S. Dollar by the RBI" in i: 
            s = i[i.find('href="')+6:]
            s = s[:s.find('">')]
            if "http" not in s:
                s = "https://www.rbi.org.in/Scripts/" + s
                bulletin_4.append(s)

        if "4A. Maturity Breakdown (by Residual Maturity) of Outstanding Forwards of RBI" in i:
            s = i[i.find('href="')+6:]
            s = s[:s.find('">')]
            if "http" not in s:
                s = "https://www.rbi.org.in/Scripts/" + s
                bulletin_4A.append(s)

        if "31. Foreign Trade" in i:
            s = i[i.find('href="')+6:]
            s = s[:s.find('">')]
            if "http" not in s:
                s = "https://www.rbi.org.in/Scripts/" + s
                bulletin_31.append(s)
                
        if "34. Foreign Investment Inflows" in i:
            s = i[i.find('href="')+6:]
            s = s[:s.find('">')]
            if "http" not in s:
                s = "https://www.rbi.org.in/Scripts/" + s
                bulletin_34.append(s)

        if "37. External Commercial Borrowings (ECBs) - Registrations" in i:
            s = i[i.find('href="')+6:]
            s = s[:s.find('">')]
            if "http" not in s:
                s = "https://www.rbi.org.in/Scripts/" + s
                bulletin_37.append(s)

        if "35. Outward Remittances under the Liberalised Remittance Scheme (LRS) for Resident Individuals" in i:
            s = i[i.find('href="')+6:]
            s = s[:s.find('">')]
            if "http" not in s:
                s = "https://www.rbi.org.in/Scripts/" + s
                bulletin_35.append(s)

        if "36. Indices of Nominal Effective Exchange Rate (NEER)" in i:
            s = i[i.find('href="')+6:]
            s = s[:s.find('">')]
            if "http" not in s:
                s = "https://www.rbi.org.in/Scripts/" + s
                bulletin_36.append(s)

        if "Indices of Real Effective Exchange Rate (REER) and Nominal Effective Exchange Rate (NEER) of the Indian Rupee" in i:
            s = i[i.find('href="')+6:]
            s = s[:s.find('">')]
            if "http" not in s:
                s = "https://www.rbi.org.in/Scripts/" + s
                bulletin_36.append(s)


    return bulletin_4,bulletin_4A,bulletin_31,bulletin_34,bulletin_35,bulletin_36,bulletin_37


def bull_36(bulletin_36):
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    d=[]
    tr=[]
    tn=[]
    en=[]
    er=[]
    
    for url in bulletin_36:
        results = requests.get(url, headers=headers) 
        soup = BeautifulSoup(results.text, "html.parser")
        tables = soup.find_all('table', class_="tablebg")
        
        year = (tables[1].find_all("tr")[0].find_all("td")[4]).get_text()
        month = ((tables[1].find_all("tr")[1].find_all("td")[2]).get_text()).replace(".","")
        try:
            date = datetime.strptime(month + " " + year, "%B %Y")
        except:
            date = datetime.strptime(month + " " + year, " %B %Y")

        d.append(date)

        Trade_NEER = (tables[1].find_all("tr")[5].find_all("td")[5]).get_text()
        tn.append(Trade_NEER)
        Trade_REER = (tables[1].find_all("tr")[6].find_all("td")[5]).get_text()
        tr.append(Trade_REER)
        Export_NEER = (tables[1].find_all("tr")[8].find_all("td")[5]).get_text()
        en.append(Export_NEER)
        Export_REER = (tables[1].find_all("tr")[9].find_all("td")[5]).get_text()
        er.append(Export_REER)

    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')#change the link to your excel file
    sheet = wb.worksheets[1]
    for i in range(len(d)):
        date = d[i]
        for cell in sheet['A']:
            if cell.value == 'Date' or cell.value ==None: 
                continue
            j = cell.value 
            #print(j)
            x = datetime.strptime(j, "%d-%m-%Y")
            if x.month == date.month and x.year==date.year:
                sheet['AG'+str(cell.row)] = float(tn[i])
                sheet['AH'+str(cell.row)] = float(tr[i])
                sheet['AI'+str(cell.row)] = float(en[i])
                sheet['AJ'+str(cell.row)] = float(er[i]) 

    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')


    return d,tn,tr,en,er

def bull_35(bulletin_35):
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    d = []
    otr =[]
    for url in bulletin_35:
        results = requests.get(url, headers=headers) 
        soup = BeautifulSoup(results.text, "html.parser")
        tables = soup.find_all('table', class_="tablebg")
        
        year = (tables[1].find_all("tr")[1].find_all("td")[3]).get_text()
        month = ((tables[1].find_all("tr")[2].find_all("td")[3]).get_text()).replace(".","")
        check= month + " " + year
        print(check)
        date = datetime.strptime(month + " " + year, "%b %Y") 
        d.append(date)
        
        outwardRemittances = (tables[1].find_all("tr")[4].find_all("td")[5]).get_text()
        otr.append(float(outwardRemittances))
    
    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')#change the link to your excel file
    sheet = wb.worksheets[1]
    for i in range(len(d)):
        date = d[i]
        for cell in sheet['A']:
            if cell.value == 'Date' or cell.value ==None: 
                continue
            j = cell.value 
            print(j)
            x = datetime.strptime(j, "%d-%m-%Y")
            if x.month == date.month and x.year==date.year:
                sheet['AK'+str(cell.row)] = otr[i]

    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')

    return d,otr

def bull_34(bulletin_34):
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    d =[]
    fdi =[]
    for url in bulletin_34:
        results = requests.get(url, headers=headers) 
        soup = BeautifulSoup(results.text, "html.parser")
        tables = soup.find_all('table', class_="tablebg")
        
        year = (tables[1].find_all("tr")[1].find_all("td")[5]).get_text()
        month = ((tables[1].find_all("tr")[2].find_all("td")[4]).get_text()).replace(".","")
        date = datetime.strptime(month + " " + year, "%b %Y")
        d.append(date) 
        
        netFDI = (tables[1].find_all("tr")[4].find_all("td")[6]).get_text()
        netFDI = netFDI.replace(",","")
        netFDI = float(netFDI)
        fdi.append(netFDI)


    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    sheet = wb.worksheets[1]
    for i in range(len(d)):
        date = d[i]
        for cell in sheet['A']:
            if cell.value == 'Date' or cell.value ==None: 
                continue
            j = cell.value 
            print(j)
            x = datetime.strptime(j, "%d-%m-%Y")
            if x.month == 6 and x.year == 2020:
                sheet['AL'+str(cell.row)] = -838
            if x.month == date.month and x.year==date.year:
                sheet['AL'+str(cell.row)] = fdi[i]
            

    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')


    return d,fdi

def bull_31(bulletin_31):
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    d =[]
    tb =[]
    tboil =[]
    tbnonoil = []
    for url in bulletin_31:
        results = requests.get(url, headers=headers) 
        soup = BeautifulSoup(results.text, "html.parser")
        tables = soup.find_all('table', class_="tablebg")
        
        year = (tables[1].find_all("tr")[0].find_all("td")[4]).get_text()
        month = ((tables[1].find_all("tr")[1].find_all("td")[5]).get_text()).replace(".","")
        date = datetime.strptime(month + " " + year, "%b %Y") 

        print(date)
        d.append(date)
        
        TradeBalance = (tables[1].find_all("tr")[16].find_all("td")[8]).get_text()
        tb.append(TradeBalance)
        TBOil = (tables[1].find_all("tr")[18].find_all("td")[8]).get_text()
        tboil.append(TBOil)
        TBNonOil = (tables[1].find_all("tr")[20].find_all("td")[8]).get_text()
        tbnonoil.append(TBNonOil)

    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    sheet = wb.worksheets[1]
    for i in range(len(d)):
        date = d[i]
        for cell in sheet['A']:
            if cell.value == 'Date' or cell.value ==None: 
                continue
            j = cell.value 
            #print(j)
            x = datetime.strptime(j, "%d-%m-%Y")
            if x.month == date.month and x.year==date.year:
                sheet['AM'+str(cell.row)] = float(tb[i])
                sheet['AN'+str(cell.row)] = float(tboil[i])
                sheet['AO'+str(cell.row)] = float(tbnonoil[i])
            

    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')
    
    return tb,tboil,tbnonoil

def bull4a(bulletin_4A):
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    d = []
    onemonth =[]
    onetothree =[]
    threetoone = []
    morethanone = []
    total =[]

    for url in bulletin_4A:
        print(url)
        results = requests.get(url, headers=headers) 
        soup = BeautifulSoup(results.text, "html.parser")
        tables = soup.find_all('table', class_="tablebg")
        date = tables[0].text
        date = " ".join(date.strip().split())
        date = date.replace(",","")
        date = date[date.find("As on")+6:]
        date = datetime.strptime(" ".join(date.split(" ")[:3]), "%B %d %Y")
        print(date)
        d.append(date)
        data = tables[0].text
        data = " ".join(data.strip().split())
        uptoOneMonth = data[data.find(" Upto 1 month")+14:]
        uptoOneMonth = uptoOneMonth.split()
        uptoOneMonth = float(uptoOneMonth[2])
        onemonth.append(uptoOneMonth)
        OneToThreeMonths = data[data.find("More than 1 month and upto 3 months")+36:]
        OneToThreeMonths = OneToThreeMonths.split()
        OneToThreeMonths = float(OneToThreeMonths[2])
        onetothree.append(OneToThreeMonths)
        ThreeMonthsToOneYear = data[data.find("More than 3 months and upto 1 year")+35:]
        ThreeMonthsToOneYear = ThreeMonthsToOneYear.split()
        ThreeMonthsToOneYear = float(ThreeMonthsToOneYear[2])
        threetoone.append(ThreeMonthsToOneYear)
        MoreThanAYear = data[data.find("More than 1 year")+17:]
        MoreThanAYear = MoreThanAYear.split()
        MoreThanAYear = float(MoreThanAYear[2])
        morethanone.append(MoreThanAYear)
        Total = (tables[1].find_all("tr")[7].find_all("td")[3]).get_text()
        Total = data[data.find("Total (1+2+3+4)")+16:]
        Total = Total.split()
        Total = float(Total[2])
        total.append(Total)

    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    sheet = wb.worksheets[1]
    for i in range(len(d)):
        date = d[i]
        for cell in sheet['A']:
            if cell.value == 'Date' or cell.value ==None: 
                continue
            j = cell.value 
            #print(j)
            x = datetime.strptime(j, "%d-%m-%Y")
            if x.month == date.month and x.year==date.year:
                sheet['AP'+str(cell.row)] = float(onemonth[i])
                sheet['AQ'+str(cell.row)] = float(onetothree[i])
                sheet['AR'+str(cell.row)] = float(threetoone[i])
                sheet['AS'+str(cell.row)] = float(morethanone[i])
                sheet['AT'+str(cell.row)] = float(total[i])

            

    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')

    return d,onemonth,onetothree,threetoone,morethanone,total


def bull4(bulletin_4):
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    d =[]
    p =[]
    s =[]
    n =[]
    for url in bulletin_4:
        results = requests.get(url, headers=headers) 
        soup = BeautifulSoup(results.text, "html.parser")
        tables = soup.find_all('table', class_="tablebg")
        
        year = (tables[1].find_all("tr")[1].find_all("td")[3]).get_text()
        month = ((tables[1].find_all("tr")[2].find_all("td")[2]).get_text()).replace(".","")
        date = datetime.strptime(month + " " + year, "%b %Y")
        d.append(date) 
        purchase = (tables[1].find_all("tr")[5].find_all("td")[4]).get_text()
        p.append(purchase)
        sale = (tables[1].find_all("tr")[6].find_all("td")[4]).get_text()
        s.append(sale)
        net = (tables[1].find_all("tr")[4].find_all("td")[4]).get_text()
        n.append(net)
    
    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    sheet = wb.worksheets[1]
    for i in range(len(d)):
        date = d[i]
        for cell in sheet['A']:
            if cell.value == 'Date' or cell.value ==None: 
                continue
            j = cell.value 
            print(j)
            x = datetime.strptime(j, "%d-%m-%Y")
            if x.month == date.month and x.year==date.year:
                sheet['AU'+str(cell.row)] = float(p[i])
                sheet['AV'+str(cell.row)] = float(s[i])
                sheet['AW'+str(cell.row)] = float(n[i])      

    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')

    return d,p,s,n

def extract(**kwargs):
    b4=[]
    b4a =[]
    b31 =[]
    b34 =[]
    b35=[]
    b36=[]
    b37=[]

    headers = {"Accept-Language": "en-US, en;q=0.5"}
    url = "https://www.rbi.org.in/Scripts/BS_ViewBulletin.aspx"
    driver = webdriver.Chrome('/usr/local/bin/chromedriver',options=chrome_options)
    driver.implicitly_wait(30)
    driver.get(url)
    strs = []
    for a in range(2020,2023):
        for b in range(1,13):
            id1 = str(a)+str(b)
            id2 = "document.getElementById('"+id1+"').click()"
            strs.append(id2)
    for i in strs:
        driver.execute_script(i)
        elem = driver.find_element(By.CLASS_NAME,"tablebg")
        bulletin_4,bulletin_4A,bulletin_31,bulletin_34,bulletin_35,bulletin_36,bulletin_37 = link_gen(elem)
        b4+= bulletin_4
        b4a+= bulletin_4A
        b31+= bulletin_31
        b34+= bulletin_34
        b35+= bulletin_35
        b36+= bulletin_36
        b37+= bulletin_37

    return b4,b4a,b31,b34,b35,b36,b37


def load(**context):
    b4 = list(context['ti'].xcom_pull(task_ids='extract')[0])
    b4a = list(context['ti'].xcom_pull(task_ids='extract')[1])
    b31 = list(context['ti'].xcom_pull(task_ids='extract')[2])
    b34 = list(context['ti'].xcom_pull(task_ids='extract')[3])
    b35 = list(context['ti'].xcom_pull(task_ids='extract')[4])
    b36 = list(context['ti'].xcom_pull(task_ids='extract')[5])
    b37 = list(context['ti'].xcom_pull(task_ids='extract')[6])

    d,tn,tr,en,er = bull_36(b36)
    d,otr = bull_35(b35)
    d,fdi = bull_34(b34)
    tb,tboil,tbnonoil = bull_31(b31)
    d,onemonth,onetothree,threetoone,morethanone,total = bull4a(b4a)
    d,p,s,n = bull4(b4)

    return

def extract2(**kwargs):
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    url = "https://rbi.org.in/Scripts/Pr_DataRelease.aspx?SectionID=352&DateFilter=Year"
    driver = webdriver.Chrome('/usr/local/bin/chromedriver',options=chrome_options)
    driver.implicitly_wait(30)
    driver.get(url)
    strs = ["document.getElementById('2020').click()","document.getElementById('2021').click()","document.getElementById('2022').click()"]
    link_arr = []    
    for k in strs:
        print(k)
        driver.execute_script(k)
        elem = driver.find_element(By.CLASS_NAME,"tablebg")
        soup_level3=BeautifulSoup(driver.page_source, 'lxml')
        zz = soup_level3.find_all("a",{"class": "link2"})
        for i in zz:
            x = str(i)
            x = x.replace("<"," ")
            x = x.replace(">"," ")
            x = x.replace('"','')
            x = " ".join(x.strip().split())
            x = x[x.find("href=")+5:].split(" ")
            x = "https://rbi.org.in/Scripts/"+x[0]
            link_arr.append(x)

    return link_arr

def load2(**context):
    urls = list(context['ti'].xcom_pull(task_ids='extract2'))
    imp =[]
    exp =[]
    d =[]
    headers = {"Accept-Language": "en-US, en;q=0.5"}
    for url in urls:
        print(url)
        results = requests.get(url, headers=headers) 
        soup = BeautifulSoup(results.text, "html.parser")
        tables = soup.find_all('table', class_="tablebg")
        data = tables[0].text
        data = data.replace("-","")
        data = data.replace("–","")
        data= " ".join(data.strip().split())
        data = re.sub("\(.*?\)","",data)
        data= " ".join(data.strip().split())
        str1 = "Notes:"
        str2 = "Note"
        if str1 in data:
            data = data[data.find("Payments")+9:data.find("Notes")].split()
            print(data)
            imports = float(data[-1].replace(",",""))
            imp.append(imports)
            exports = float(data[-2].replace(",",""))
            exp.append(exports)
            try:
                date = data[-4]+" "+data[-3]
                date = datetime.strptime(date,"%B %Y")
                d.append(date)
            except:
                date = data[-3]
                d1 = date[:date.find("2")]
                d2 = date[date.find("2"):]
                date = d1+" "+d2
                date = datetime.strptime(date,"%B %Y")
                d.append(date)
        else:
            data = data[data.find("Payments")+9:data.find("Note")].split()
            print(data)
            imports = float(data[-1].replace(",",""))
            imp.append(imports)
            exports = float(data[-2].replace(",",""))
            exp.append(exports)
            try:
                date = data[-4]+" "+data[-3]
                date = datetime.strptime(date,"%B %Y")
                d.append(date)
            except:
                date = data[-3]
                d1 = date[:date.find("2")]
                d2 = date[date.find("2"):]
                date = d1+" "+d2
                date = datetime.strptime(date,"%B %Y")
                d.append(date)

    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    sheet = wb.worksheets[1]
    for i in range(len(d)):
        date = d[i]
        for cell in sheet['A']:
            if cell.value == 'Date' or cell.value ==None: 
                continue
            j = cell.value 
            print(j)
            x = datetime.strptime(j, "%d-%m-%Y")
            if x.month == date.month and x.year==date.year:
                sheet['AX'+str(cell.row)] = imp[i]
                sheet['AY'+str(cell.row)] = exp[i]

    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')

    return 



default_args = {
    'owner': 'airflow',
    'depends_on_past': False,
    'start_date': datetime(2023,1,4),
    'email': ['nikheleshbhattacharyya@gmail.com'],
    'email_on_failure': True,
    'email_on_retry': False,
    'retries': 3,
    'retry_delay': timedelta(minutes=2)
}

schedule_interval = '0 */2 * * *'

dag = DAG(
    dag_id = 'Monthly',
    default_args = default_args,
    schedule = '@daily'
)

task_start = BashOperator(
    task_id = 'start_task',
    bash_command = 'echo start',
    dag = dag
)

extract_data = PythonOperator(
    task_id = 'extract',
    python_callable = extract,
    provide_context = True,
    dag = dag
)

load_data = PythonOperator(
    task_id = 'load',
    python_callable = load,
    provide_context = True,
    dag = dag
)

extract_data2 = PythonOperator(
    task_id = 'extract2',
    python_callable = extract2,
    provide_context = True,
    dag = dag
)

load_data = PythonOperator(
    task_id = 'load2',
    python_callable = load2,
    provide_context = True,
    dag = dag
)


#send_email = EmailOperator(
#    task_id = 'send_email',
#    to = ['nikheleshbhattacharyya@gmail.com'],
#    subject = 'Covid19 data ',
#    html_content = '''
#    {date} Data data has been scraped!
#    '''.format(date = datetime.now()),
#    dag = dag
#)

finish_start = BashOperator(
    task_id = 'finish_task',
    bash_command = 'echo finish',
    dag = dag
)


task_start>>extract_data>>load_data>>finish_start